import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# def highlight_pdf_rows(folder_path, excel_path, target_column='A', check_subfolders=True, with_extension=False, case_sensitive=False):
#     """
#     将包含PDF文件名的Excel行标黄
    
#     :param folder_path: PDF文件夹路径
#     :param excel_path: Excel文件路径
#     :param target_column: 目标列（字母或数字）
#     :param check_subfolders: 是否检查子文件夹
#     :param with_extension: 是否包含扩展名
#     :param case_sensitive: 是否区分大小写
#     """
#     # 收集所有PDF文件名
#     pdf_names = set()
    
#     if check_subfolders:
#         for root, _, files in os.walk(folder_path):
#             for file in files:
#                 if file.lower().endswith('.pdf'):
#                     name = file if with_extension else os.path.splitext(file)[0]
#                     pdf_names.add(name if case_sensitive else name.lower())
#     else:
#         for file in os.listdir(folder_path):
#             if file.lower().endswith('.pdf'):
#                 name = file if with_extension else os.path.splitext(file)[0]
#                 pdf_names.add(name if case_sensitive else name.lower())

#     # 加载Excel文件
#     wb = load_workbook(excel_path)
#     ws = wb.active
#     yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

#     # 转换列标识
#     if isinstance(target_column, str):
#         col_idx = ord(target_column.upper()) - ord('A')
#     else:
#         col_idx = target_column - 1

#     # 遍历Excel行
#     for row in ws.iter_rows():
#         cell = row[col_idx]
#         if cell.value is None:
#             continue
        
#         # 处理比较值
#         compare_value = str(cell.value)
#         if not case_sensitive:
#             compare_value = compare_value.lower()
        
#         if compare_value in pdf_names:
#             for cell_in_row in row:
#                 cell_in_row.fill = yellow_fill

#     wb.save(excel_path)

def highlight_pdf_rows_multiple(folder_path, excel_paths, target_column='A', check_subfolders=True, with_extension=False, case_sensitive=False):
    """
    将包含PDF文件名的Excel行标黄，并输出未匹配的PDF文件名
    
    :param folder_path: PDF文件夹路径
    :param excel_paths: Excel文件路径列表
    :param target_column: 目标列（字母或数字）
    :param check_subfolders: 是否检查子文件夹
    :param with_extension: 是否包含扩展名
    :param case_sensitive: 是否区分大小写
    """
    # 收集所有PDF文件名
    pdf_names = set()
    if check_subfolders:
        for root, _, files in os.walk(folder_path):
            for file in files:
                if file.lower().endswith('.pdf'):
                    name = file if with_extension else os.path.splitext(file)[0]
                    pdf_names.add(name if case_sensitive else name.lower())
    else:
        for file in os.listdir(folder_path):
            if file.lower().endswith('.pdf'):
                name = file if with_extension else os.path.splitext(file)[0]
                pdf_names.add(name if case_sensitive else name.lower())

    unmatched_pdfs = pdf_names.copy()  # 用于记录未匹配的PDF文件名
    green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')

    for excel_path in excel_paths:
        # 加载Excel文件
        wb = load_workbook(excel_path)
        ws = wb.active

        # 转换列标识
        if isinstance(target_column, str):
            col_idx = ord(target_column.upper()) - ord('A')
        else:
            col_idx = target_column - 1

        # 遍历Excel行
        for row in ws.iter_rows():
            cell = row[col_idx]
            if cell.value is None:
                continue
            
            # 处理比较值
            compare_value = str(cell.value)
            if not case_sensitive:
                compare_value = compare_value.lower()
            
            if compare_value in unmatched_pdfs:
                unmatched_pdfs.remove(compare_value)  # 从未匹配列表中移除
                for cell_in_row in row:
                    cell_in_row.fill = green_fill

        wb.save(excel_path)

    # 输出未匹配的PDF文件名
    if unmatched_pdfs:
        print("未匹配的PDF文件名:")
        for pdf in unmatched_pdfs:
            print(pdf)

# # 使用示例
# highlight_pdf_rows(
#     folder_path='/home/shy/202410',
#     excel_path='/home/shy/1.xlsx',
#     target_column='B',          # 目标（支持字母或数字）
#     check_subfolders=True,      # 包含子文件夹
#     with_extension=False,       # 不比较扩展名
#     case_sensitive=False        # 不区分大小写
# )

# 使用示例
highlight_pdf_rows_multiple(
    folder_path='/home/shy/图片/财务凭证_new/PDF_lyl',  # PDF文件夹路径
    excel_paths=['/home/shy/1.xlsx', '/home/shy/2.xlsx'],  # 两个Excel文件路径
    target_column='B',          # 目标列（支持字母或数字）
    check_subfolders=True,      # 包含子文件夹
    with_extension=False,       # 不比较扩展名
    case_sensitive=False        # 不区分大小写
)


