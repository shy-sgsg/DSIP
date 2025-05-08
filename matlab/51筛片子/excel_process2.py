from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def highlight_cells(filename):
    # 加载工作簿
    wb = load_workbook(filename)
    
    for sheet in wb.worksheets:
        # 初始化各列的范围字典，键为列索引（2到7对应C到H列）
        column_ranges = {col: {'min': None, 'max': None} for col in range(2, 10)}
        
        # 第一遍遍历：收集各列的最小值和最大值
        for row in sheet.iter_rows(min_row=2):  # 假设标题行在第一行
            for col_idx in range(2, 10):  # 处理C到H列
                cell = row[col_idx]
                if col_idx == 6 or col_idx == 7:
                    column_ranges[col_idx]['min'] = 40
                    column_ranges[col_idx]['max'] = 80
                elif col_idx == 8 or col_idx == 9:
                    column_ranges[col_idx]['min'] = 75
                    column_ranges[col_idx]['max'] = 90
                else:
                    column_ranges[col_idx]['min'] = 30
                    column_ranges[col_idx]['max'] = 55

                # if isinstance(cell.value, (int, float)):
                #     value = cell.value
                #     # 更新最小值
                #     if column_ranges[col_idx]['min'] is None or value < column_ranges[col_idx]['min']:
                #         column_ranges[col_idx]['min'] = value
                #     # 更新最大值
                #     if column_ranges[col_idx]['max'] is None or value > column_ranges[col_idx]['max']:
                #         column_ranges[col_idx]['max'] = value
        
        # 第二遍遍历：应用渐变色
        for row in sheet.iter_rows(min_row=2):
            for col_idx in range(2, 10):
                cell = row[col_idx]
                if isinstance(cell.value, (int, float)):
                    col_min = column_ranges[col_idx]['min']
                    col_max = column_ranges[col_idx]['max']
                    
                    # 如果列中没有有效数据，跳过
                    if col_min is None or col_max is None:
                        continue
                    
                    # 计算归一化比例（0到1之间）
                    if col_max == col_min:
                        p = 1.0  # 所有值相同，设为最大值颜色（绿色）
                    else:
                        p = (cell.value - col_min) / (col_max - col_min)
                        p = max(0.0, min(1.0, p))  # 确保p在0-1之间
                    
                    # 计算红绿分量（红递减，绿递增）
                    red = int(255 * (1 - p))
                    green = int(255 * p)
                    blue = 0
                    
                    # 生成十六进制颜色代码
                    color_code = f"{red:02X}{green:02X}{blue:02X}"
                    
                    # 创建填充样式
                    gradient_fill = PatternFill(
                        start_color=color_code,
                        end_color=color_code,
                        fill_type='solid'
                    )
                    cell.fill = gradient_fill
    
    # 保存为新文件
    new_filename = filename.replace('.xlsx', '_gradient.xlsx')
    wb.save(new_filename)
    print(f"处理完成！已保存为: {new_filename}")

# 使用示例
highlight_cells("SFDR_NoiseFloor_Report.xlsx")