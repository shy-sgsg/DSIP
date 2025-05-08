from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def highlight_cells(filename):
    green_fill = PatternFill(start_color='00FF00',
                             end_color='00FF00',
                             fill_type='solid')
    red_fill = PatternFill(start_color='FF0000',
                             end_color='FF0000',
                             fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFF00',
                             end_color='FFFF00',
                             fill_type='solid')
    orange_fill = PatternFill(start_color='FF8C00',
                             end_color='FF8C00',
                             fill_type='solid')
    
    # 加载工作簿
    wb = load_workbook(filename)
    
    # 遍历所有工作表 (按需修改)
    for sheet in wb.worksheets:
        # 遍历数据行 (假设数据从第2行开始)
        for row in sheet.iter_rows(min_row=2):  # 如果包含标题行则从2开始
            # 处理第三列 (索引2)
            if row[2].value and isinstance(row[2].value, (int, float)):
                if row[2].value < 40:
                    row[2].fill = red_fill
                if row[2].value >= 40 and row[2].value < 45:
                    row[2].fill = orange_fill
                if row[2].value >= 45 and row[2].value < 50:
                    row[2].fill = yellow_fill
                if row[2].value >= 50:
                    row[2].fill = green_fill  
                    
            # 处理第四列 (索引3)
            if row[3].value and isinstance(row[3].value, (int, float)):
                if row[3].value < 40:
                    row[3].fill = red_fill
                if row[3].value >= 40 and row[3].value < 45:
                    row[3].fill = orange_fill
                if row[3].value >= 45 and row[3].value < 50:
                    row[3].fill = yellow_fill
                if row[3].value >= 50:
                    row[3].fill = green_fill  

            # 处理第五列 (索引4)
            if row[4].value and isinstance(row[4].value, (int, float)):
                if row[4].value < 60:
                    row[4].fill = red_fill
                if row[4].value >= 60 and row[4].value < 70:
                    row[4].fill = orange_fill
                if row[4].value >= 70 and row[4].value < 75:
                    row[4].fill = yellow_fill
                if row[4].value >= 75:
                    row[4].fill = green_fill
            
            # 处理第六列 (索引5)        
            if row[5].value and isinstance(row[5].value, (int, float)):
                if row[5].value < 60:
                    row[5].fill = red_fill
                if row[5].value >= 60 and row[5].value < 70:
                    row[5].fill = orange_fill
                if row[5].value >= 70 and row[5].value < 75:
                    row[5].fill = yellow_fill
                if row[5].value >= 75:
                    row[5].fill = green_fill

            # 处理第七列 (索引6)
            if row[6].value and isinstance(row[6].value, (int, float)):
                if row[6].value < 40:
                    row[6].fill = red_fill
                if row[6].value >= 40 and row[6].value < 45:
                    row[6].fill = orange_fill
                if row[6].value >= 45 and row[6].value < 50:
                    row[6].fill = yellow_fill
                if row[6].value >= 50:
                    row[6].fill = green_fill

            # 处理第八列 (索引7)    
            if row[7].value and isinstance(row[7].value, (int, float)):
                if row[7].value < 40:
                    row[7].fill = red_fill
                if row[7].value >= 40 and row[7].value < 45:
                    row[7].fill = orange_fill
                if row[7].value >= 45 and row[7].value < 50:
                    row[7].fill = yellow_fill
                if row[7].value >= 50:
                    row[7].fill = green_fill

            
   

    # 保存修改 (建议使用新文件名)
    new_filename = filename.replace('.xlsx', '_highlighted.xlsx')
    wb.save(new_filename)
    print(f"处理完成！已保存为: {new_filename}")

# 使用示例
highlight_cells("SFDR_Report.xlsx")  # 替换为实际文件名